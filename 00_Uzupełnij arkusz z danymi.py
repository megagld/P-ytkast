from bs4 import BeautifulSoup
import requests
import regex as re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def pobierz_liste_odc():
    # pobierz listę gotowych odcinków z arkusza

    global gotowe_odc

    # Plik z danymi:
    file = '{}{}'.format(Path(__file__).parent,'''/Płytkast.xlsx''')

    # Nazwy arkuszy:
    wb=load_workbook(file)
    ws = wb.sheetnames

    for i in ws:
        print(i,'- gotowe')

    gotowe_odc=ws

def pobierz_dane_ze_strony():

    # pobieranie listy odcinków wraz z linkami do stron

    global odcinki

    lista_odcinkow='https://podcasts.google.com/feed/aHR0cHM6Ly9hbmNob3IuZm0vcy84NTU2OWMyOC9wb2RjYXN0L3Jzcw?sa=X&ved=0CAMQ4aUDahcKEwioq4Ca2pH5AhUAAAAAHQAAAAAQAg'

    s_get = requests.get(lista_odcinkow)

    soup = BeautifulSoup(s_get.content, "html.parser")

    chap_pat='([\d]{2}:[\d]{2}:[\d]{2}.*)'

    part_nrs =   [i.text.split(': ')[0] for i in soup.find_all("div", class_="e3ZUqe")]
    titles =     [i.text.split(': ')[1] for i in soup.find_all("div", class_="e3ZUqe")]
    main_urls =  ['https://podcasts.google.com'+i['href'] for i in soup.find_all("a", class_="D9uPgd")]
    mp3_urls =   [i['jsdata'].split(';')[1] for i in soup.find_all("div", jsname="fvi9Ef")]
    chapters =   [re.findall(chap_pat, i.text) for i in soup.find_all("div", class_="LrApYe")]

    odcinki=zip(part_nrs,titles,main_urls,mp3_urls,chapters)

def znajdz_brakujace():
    # porównuje listę odcinków z arkusza z tą ze strony
    global odcinki, gotowe_odc

    for part_nr,title,main_url,mp3_url,chapters in odcinki:
        if part_nr not in gotowe_odc:
            dodaj_odcinek(part_nr,title,main_url,mp3_url,chapters)

def dodaj_odcinek(part_nr,title,main_url,mp3_url,chapters):
    # dodaje arkusz z danymi z odcinka
    file = '{}{}'.format(Path(__file__).parent,'''/Płytkast.xlsx''')
    wb=load_workbook(file)
    wb.create_sheet(part_nr)
    ws = wb[part_nr]
    for i,j in enumerate(chapters):
        ws.cell(row=i+4,column=1).value=i+1
        ws.cell(i+4, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i+4,column=2).value=j
    ws.cell(row=2,column=2).value=title
    ws.cell(row=2,column=4).value=main_url
    ws.cell(row=3,column=4).value=mp3_url

    ws.column_dimensions['A'].width=3
    ws.column_dimensions['B'].width=30

    wb.save(file)
    wb.close()

    print('xxxxxxxxxxxxxxxxxxxxxxxxx\n')
    print('Dodano {}: {} do arkusza. Poniżej link do podcastu.\n'.format(part_nr,title))
    print(mp3_url,'\n')
    print('xxxxxxxxxxxxxxxxxxxxxxxxx')


def r():    
    pobierz_liste_odc()

    pobierz_dane_ze_strony()

    znajdz_brakujace()
